import streamlit as st
import pandas as pd
from io import BytesIO
import gender_guesser.detector as gender
from openpyxl.utils import get_column_letter
import os
import unicodedata
import re
import streamlit.components.v1 as components

# --- INICIALIZA O MOTOR OFFLINE ---
@st.cache_resource
def carregar_motor_offline():
    return gender.Detector(case_sensitive=False)

detector = carregar_motor_offline()

# --- FUNÇÕES DE BUSCA DE COLUNAS ---
def encontrar_coluna_nome(columns):
    for col in columns:
        col_clean = str(col).strip().upper()
        if col_clean in ["NOME", "NOME COMPLETO", "NOME_COMPLETO", "NOME CLIENTE", "CLIENTE"]:
            return col
    for col in columns:
        if "NOME" in str(col).strip().upper():
            return col
    return None

def encontrar_coluna_cpf(columns):
    for col in columns:
        col_clean = str(col).strip().upper()
        if col_clean in ["CPF", "CPF/CNPJ", "CPF_CNPJ", "DOCUMENTO", "DOC"]:
            return col
    for col in columns:
        if "CPF" in str(col).strip().upper():
            return col
    return None

# --- FUNÇÃO DE HIGIENIZAÇÃO DE NOMES (MAIÚSCULAS) ---
def limpar_nome(nome):
    if not nome or pd.isna(nome):
        return ""
    
    # 1. Remove acentos e cedilhas
    nome_normalizado = unicodedata.normalize('NFD', str(nome))
    nome_sem_acentos = ''.join(c for c in nome_normalizado if unicodedata.category(c) != 'Mn')
    
    # 2. Remove pontuações, traços e caracteres especiais (mantém apenas letras e espaços)
    nome_limpo = re.sub(r'[^a-zA-Z\s]', '', nome_sem_acentos)
    
    # 3. Remove espaços duplos e converte para MAIÚSCULAS
    return ' '.join(nome_limpo.split()).upper()

# --- FUNÇÃO DE LIMPEZA DE CPF (SEM PONTOS, TRAÇOS E SEM ZEROS À ESQUERDA) ---
def limpar_cpf(cpf):
    if pd.isna(cpf) or cpf is None:
        return ""
    
    # Converte para string e remove decimais do pandas (.0)
    cpf_str = str(cpf).split('.')[0].strip()
    
    # Remove pontos e traços
    cpf_limpo = re.sub(r'[\.\-]', '', cpf_str)
    
    # Remove zeros à esquerda e converte para valor limpo
    cpf_sem_zero = cpf_limpo.lstrip('0')
    
    return cpf_sem_zero

# --- FUNÇÃO DE CLASSIFICAÇÃO DE GÊNERO ---
def classificar_genero_rapido(nome_completo):
    if not nome_completo or pd.isna(nome_completo):
        return ""
    
    nome_limpo = limpar_nome(nome_completo)
    if not nome_limpo:
        return ""
        
    primeiro_nome = nome_limpo.split()[0].title()
    resultado = detector.get_gender(primeiro_nome)
    
    if resultado in ['male', 'mostly_male']:
        return "M"
    elif resultado in ['female', 'mostly_female']:
        return "F"
    else:
        nome_min = primeiro_nome.lower()
        if nome_min.endswith(('a', 'z', 'y', 'elly', 'ine', 'ane', 'ele', 'ia', 'ce', 'te', 'is')):
            return "F"
        else:
            return "M"

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="Classificador - SulAmérica", 
    page_icon="logo.png" if os.path.exists("logo.png") else "💙", 
    layout="centered"
)

# --- ESTILO CUSTOMIZADO ---
st.markdown("""
    <style>
    /* Fundo Gradiente Azul SulAmérica */
    .stApp {
        background: linear-gradient(135deg, #001A3B 0%, #002D62 50%, #0B4B8A 100%);
        background-attachment: fixed;
    }

    /* FORÇA TODOS OS TEXTOS PADRÃO E RÓTULOS PARA BRANCO */
    html, body, p, span, label, h1, h2, h3, h4, .stMarkdown {
        color: #FFFFFF !important;
        font-size: 18px;
    }
    
    .header-sub {
        color: #E2E8F0 !important;
        font-size: 18px !important;
        margin-bottom: 20px;
    }

    .stTextInput label p, .stFileUploader label p {
        color: #FFFFFF !important;
        font-size: 18px !important;
        font-weight: bold !important;
    }

    .stTextInput input {
        font-size: 18px !important;
        padding: 12px !important;
        background-color: #FFFFFF !important;
        border: 2px solid #CBD5E1 !important;
        color: #0F172A !important;
    }

    /* TEXTO E ÍCONES DAS ABAS EM BRANCO PURO */
    button[data-baseweb="tab"] div p, 
    button[data-baseweb="tab"] p, 
    button[data-baseweb="tab"] span {
        color: #FFFFFF !important;
        font-size: 18px !important;
        font-weight: bold !important;
    }
    
    button[data-baseweb="tab"][aria-selected="true"] {
        border-bottom: 3px solid #F37021 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] div p {
        color: #F37021 !important;
    }

    /* TRADUÇÃO E ESTILO DO UPLOAD DE ARQUIVOS */
    [data-testid="stFileUploaderDropzoneInstructions"] div span {
        display: none !important;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] div::after {
        content: "Arraste e solte o arquivo aqui";
        font-size: 18px !important;
        font-weight: bold;
        color: #FFFFFF !important;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] small {
        display: none !important;
    }
    [data-testid="stFileUploaderDropzone"] {
        background-color: rgba(255, 255, 255, 0.1) !important;
        border: 2px dashed #FFFFFF !important;
    }
    [data-testid="stFileUploaderDropzone"] button {
        font-size: 0px !important;
        background-color: #F37021 !important;
        color: white !important;
        border-radius: 6px !important;
        padding: 8px 18px !important;
    }
    [data-testid="stFileUploaderDropzone"] button::after {
        content: "Procurar arquivo";
        font-size: 16px !important;
        font-weight: bold;
        display: block;
        color: #FFFFFF !important;
    }

    /* Estilo dos Botões */
    .stButton>button {
        background-color: #F37021;
        border-radius: 8px;
        padding: 12px 28px;
        border: none;
        box-shadow: 0 4px 6px rgba(243, 112, 33, 0.3);
    }
    .stButton>button div p {
        color: #FFFFFF !important;
        font-weight: bold !important;
        font-size: 18px !important;
    }
    .stButton>button:hover {
        background-color: #D95B0F;
        transform: translateY(-2px);
    }
    
    .stDownloadButton>button {
        background-color: #002D62;
        border-radius: 8px;
        padding: 12px 28px;
    }
    .stDownloadButton>button div p {
        color: #FFFFFF !important;
        font-weight: bold !important;
        font-size: 18px !important;
    }

    .header-title {
        font-size: 34px !important;
        font-weight: 800;
        color: #F37021 !important;
        margin-bottom: 5px;
    }
    </style>
""", unsafe_allow_html=True)

# Cabeçalho com Logo
col_logo, col_titulo = st.columns([1.2, 2.8])

with col_logo:
    if os.path.exists("logo.png"):
        st.image("logo.png", width=160)

with col_titulo:
    st.markdown('<p class="header-title">Classificador de Gênero</p>', unsafe_allow_html=True)
    st.markdown('<p class="header-sub">Ferramenta interna para identificação rápida por primeiro nome.</p>', unsafe_allow_html=True)

st.divider()

# --- ABAS DO SISTEMA ---
aba1, aba2, aba3 = st.tabs(["🔍 Consulta Rápida", "📁 Processamento em Lote (Excel)", "🧼 Higienização (Nome e CPF)"])

with aba1:
    st.markdown("### Consultar um único nome")
    
    col1, col2 = st.columns([2.8, 1.2])
    with col1:
        nome_digitado = st.text_input("Digite o nome do cliente:", key="input_genero_unico")
    with col2:
        st.write("")
        st.write("")
        btn_consultar = st.button("Classificar Único")
        
    if btn_consultar and nome_digitado:
        res = classificar_genero_rapido(nome_digitado)
        
        emoji = "👨‍💼 Masculino (M)" if res == "M" else "👩‍💼 Feminino (F)"
        cor_borda = "#002D62" if res == "M" else "#DB2777"
        cor_fundo = "#EFF6FF" if res == "M" else "#FDF2F8"
        cor_texto = "#1E40AF" if res == "M" else "#9D174D"
        
        st.markdown(f"""
            <div style="padding: 24px; margin-top: 18px; border-radius: 12px; background-color: {cor_fundo}; text-align: center; border: 2px solid {cor_borda}; border-left: 10px solid {cor_borda}; box-shadow: 0 4px 12px rgba(0,0,0,0.08);">
                <h3 style="margin:0; color: #0F172A !important; font-size: 28px;">{nome_digitado.upper()}</h3>
                <p style="font-size: 24px; margin: 10px 0 0 0; font-weight: 800; color: {cor_texto} !important;">{emoji}</p>
            </div>
        """, unsafe_allow_html=True)

with aba2:
    st.markdown("### Carregue seu arquivo clicando em upload")
    
    arq = st.file_uploader("", type=["xlsx"], key="uploader_genero")
    
    if arq:
        df = pd.read_excel(arq)
        col_nome = encontrar_coluna_nome(df.columns)
        
        if col_nome:
            st.write(f"📋 **Coluna identificada:** `{col_nome}`")
            st.dataframe(df.head(3), use_container_width=True)
            
            if st.button("🚀 Processar Todos os Nomes"):
                with st.spinner("Analisando nomes..."):
                    df["Gênero Identificado"] = df[col_nome].apply(classificar_genero_rapido)
                
                st.success("✅ Processamento concluído com sucesso!")
                st.snow()
                components.html("""
                    <script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1.5.1/dist/confetti.browser.min.min.js"></script>
                    <script>
                        confetti({
                            particleCount: 50,
                            spread: 60,
                            origin: { y: 0.7 },
                            colors: ['#F37021', '#002D62', '#FFFFFF']
                        });
                    </script>
                """, height=0)
                
                st.divider()
                st.write("📊 **Resumo da Classificação:**")
                
                total = len(df)
                total_m = (df["Gênero Identificado"] == "M").sum()
                total_f = (df["Gênero Identificado"] == "F").sum()
                
                colA, colB, colC = st.columns(3)
                colA.metric("Total de Clientes", total)
                colB.metric("Masculino (M)", total_m)
                colC.metric("Feminino (F)", total_f)
                
                st.divider()
                
                saida = BytesIO()
                with pd.ExcelWriter(saida, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Resultados')
                    worksheet = writer.sheets['Resultados']
                    for i, col in enumerate(df.columns):
                        tamanho_maximo = max(df[col].astype(str).map(len).max(), len(str(col)))
                        worksheet.column_dimensions[get_column_letter(i + 1)].width = tamanho_maximo + 3
                
                st.download_button("📥 Baixar Planilha Final Formatada", data=saida.getvalue(), file_name="clientes_classificados.xlsx")
        else:
            st.error("⚠️ Nenhuma coluna de nome encontrada. O cabeçalho deve ser 'Nome', 'NOME' ou variações parecidas.")

# --- ABA 3: HIGIENIZAÇÃO DE NOMES E CPF ---
with aba3:
    st.markdown("### Limpeza e Padronização (Nome e CPF)")
    st.write("Converte nomes para MAIÚSCULAS sem acentos/símbolos, remove pontos, traços e zeros à esquerda dos CPFs (Formatação Geral).")
    
    opcao_limpeza = st.radio("Escolha o modo de higienização:", ["Consulta Única", "Processar Planilha Excel"])
    
    if opcao_limpeza == "Consulta Única":
        c1, c2 = st.columns(2)
        with c1:
            nome_para_limpar = st.text_input("Nome completo:", key="input_limpar_unico")
        with c2:
            cpf_para_limpar = st.text_input("CPF:", key="input_cpf_unico")
            
        if st.button("🧼 Higienizar Dados") and (nome_para_limpar or cpf_para_limpar):
            resultado_nome = limpar_nome(nome_para_limpar) if nome_para_limpar else "-"
            resultado_cpf = limpar_cpf(cpf_para_limpar) if cpf_para_limpar else "-"
            
            st.markdown(f"""
                <div style="padding: 20px; margin-top: 15px; border-radius: 12px; background-color: #EFF6FF; border: 2px solid #002D62; border-left: 10px solid #002D62;">
                    <p style="margin:0; font-size: 18px; color: #002D62 !important;"><b>Nome Higienizado (Maiúsculo):</b> {resultado_nome}</p>
                    <p style="margin:8px 0 0 0; font-size: 18px; color: #002D62 !important;"><b>CPF Higienizado (Geral/Sem Zeros):</b> {resultado_cpf}</p>
                </div>
            """, unsafe_allow_html=True)

    else:
        arq_limpeza = st.file_uploader("Carregue seu arquivo para higienização", type=["xlsx"], key="uploader_limpeza")
        if arq_limpeza:
            df_limpeza = pd.read_excel(arq_limpeza, dtype=str)
            col_nome_limp = encontrar_coluna_nome(df_limpeza.columns)
            col_cpf_limp = encontrar_coluna_cpf(df_limpeza.columns)
            
            if col_nome_limp or col_cpf_limp:
                st.write(f"📋 **Colunas identificadas:** Nome = `{col_nome_limp}` | CPF = `{col_cpf_limp}`")
                st.dataframe(df_limpeza.head(3), use_container_width=True)
                
                if st.button("🧼 Higienizar Planilha"):
                    with st.spinner("Processando dados..."):
                        # Cria DataFrame apenas com as colunas higienizadas
                        df_final = pd.DataFrame()
                        
                        if col_nome_limp:
                            df_final["Nome"] = df_limpeza[col_nome_limp].apply(limpar_nome)
                        if col_cpf_limp:
                            df_final["CPF"] = df_limpeza[col_cpf_limp].apply(limpar_cpf)
                    
                    st.success("✅ Higienização concluída! Colunas antigas removidas.")
                    st.dataframe(df_final.head(5), use_container_width=True)
                    
                    saida_limpa = BytesIO()
                    with pd.ExcelWriter(saida_limpa, engine='openpyxl') as writer:
                        df_final.to_excel(writer, index=False, sheet_name='Higienizados')
                        worksheet = writer.sheets['Higienizados']
                        
                        for i, col in enumerate(df_final.columns):
                            tamanho_maximo = max(df_final[col].astype(str).map(len).max(), len(str(col)))
                            worksheet.column_dimensions[get_column_letter(i + 1)].width = tamanho_maximo + 3
                            
                            # Define a formatação das células para 'Geral' (General)
                            for cell in worksheet[get_column_letter(i + 1)]:
                                cell.number_format = 'General'
                    
                    st.download_button("📥 Baixar Planilha Higienizada", data=saida_limpa.getvalue(), file_name="dados_higienizados.xlsx")
            else:
                st.error("⚠️ Nenhuma coluna de Nome ou CPF foi encontrada na planilha.")
